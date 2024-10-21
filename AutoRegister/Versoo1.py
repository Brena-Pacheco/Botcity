from botcity.core import Bot  # Importa a classe base para criar bots
from botcity.web import WebBot  # Importa a classe para interagir com navegadores
from webdriver_manager.chrome import ChromeDriverManager  # Gerencia o download e configuração do ChromeDriver
from openpyxl import load_workbook  # Manipula arquivos Excel

class BotCadastroFuncionarios(WebBot):
    def action(self, execution=None):
        # Configuração do WebDriver (navegador)
        self.driver_path = ChromeDriverManager().install()  # Baixa e configura o ChromeDriver automaticamente
        self.headless = False  # Executa o navegador com interface gráfica (False) ou sem (True)

        # Abre o formulário Google Forms
        self.browse("https://docs.google.com/forms/d/e/1FAIpQLSfDzqYUoNJMY09vo4o9NNmwWWzgoNHMX3on6wctH2z3rQkB1A/viewform")

        # Abre a planilha Excel online (substitua pelo seu link)
        workbook = load_workbook(filename="https://1drv.ms/x/c/51f0ace1d6fa3427/EbSLsKeckS5PlsFkigqFZGABet9zgjSSP-jOxhEG3DEL4Q?e=6rRJ4r", read_only=False)  # Carrega a planilha, permitindo edição
        sheet = workbook.active  # Seleciona a planilha ativa (primeira por padrão)

        # Itera pelas linhas da planilha (começando da segunda linha, assumindo que a primeira é o cabeçalho)
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Lê cada linha como uma lista de valores
            nome, genero, email, departamento, endereco, cpf, rg, turno, status = row  # Desempacota os valores da linha

            # Preencha o formulário apenas se o status for "não cadastrado" (ignorando maiúsculas/minúsculas)
            if status.lower() == "não cadastrado":
                self.find_element("Nome", matching=0.7).send_keys(nome)  # Encontra o campo "Nome" e preenche com o valor

                # Seleciona o gênero (substitua pelos seletores corretos)
                if genero.lower() == "masculino":
                    self.find_element("Masculino", matching=0.7).click()  # Clica no botão de rádio "Masculino"
                else:
                    self.find_element("Feminino", matching=0.7).click()  # Clica no botão de rádio "Feminino"

                # Preenche os demais campos do formulário (substitua pelos seletores corretos)
                self.find_element("E-mail", matching=0.7).send_keys(email)
                self.find_element("Departamento", matching=0.7).send_keys(departamento)
                self.find_element("Endereço", matching=0.7).send_keys(endereco)
                self.find_element("CPF", matching=0.7).send_keys(cpf)
                self.find_element("RG", matching=0.7).send_keys(rg)

                # Seleciona o turno (substitua pelos seletores corretos)
                if turno.lower() == "primeiro":
                    self.find_element("Primeiro", matching=0.7).click()
                elif turno.lower() == "segundo":
                    self.find_element("Segundo", matching=0.7).click()
                else:
                    self.find_element("Comercial", matching=0.7).click()

                # Envie o formulário (substitua pelo seletor correto)
                self.find_element("Enviar", matching=0.7).click()

                # Atualize o status na planilha
                sheet.cell(row=row[0].row, column=9).value = "Cadastrado"  # Altera o valor da célula na coluna 9 (status)

                # Salva a planilha
                workbook.save("https://1drv.ms/x/c/51f0ace1d6fa3427/EbSLsKeckS5PlsFkigqFZGABet9zgjSSP-jOxhEG3DEL4Q?e=6rRJ4r")  # Salva as alterações na planilha online

        # Imprima uma mensagem de conclusão
        print("Cadastro de funcionários concluído!")

if __name__ == '__main__':
    BotCadastroFuncionarios.main()  # Executa o bot se o script for chamado diretamente